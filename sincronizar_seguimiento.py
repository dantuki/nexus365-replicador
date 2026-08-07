import json
import asyncio
import httpx
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import dataframe_to_rows

API_BASE_URL = "https://api.nexus-365.com/api/v1"

def load_token():
    try:
        with open("session.json", "r") as f:
            return json.load(f).get("token")
    except Exception as e:
        print(f"Error al leer session.json: {e}")
        return None

async def fetch_all_services(client, headers):
    services = []
    page = 1
    limit = 50
    
    while True:
        url = f"{API_BASE_URL}/auxservices?limit={limit}&page={page}&sort=-createdAt"
        response = await client.get(url, headers=headers)
        
        if response.status_code != 200:
            print(f"Error obteniendo página {page}: {response.status_code}")
            break
            
        payload = response.json().get("payload", {})
        docs = payload.get("docs", [])
        services.extend(docs)
        
        page_count = payload.get("totalPages", 1)
        if page >= page_count or not docs:
            break
            
        page += 1
        
    return services

async def fetch_cost_details(client, headers):
    costs_map = {}
    page = 1
    limit = 50
    
    while True:
        url = f"{API_BASE_URL}/cost/auxServices?limit={limit}&page={page}"
        response = await client.get(url, headers=headers)
        
        if response.status_code != 200:
            break
            
        payload = response.json().get("payload", {})
        docs = payload.get("docs", [])
        
        for doc in docs:
            exp_id = doc.get("auxiliaryService", {}).get("code") or doc.get("code")
            if exp_id:
                costs_map[exp_id] = doc
                
        page_count = payload.get("totalPages", 1)
        if page >= page_count or not docs:
            break
            
        page += 1
        
    return costs_map

async def generate_seguimiento_excel():
    token = load_token()
    if not token:
        print("No se encontró el token de sesión en session.json.")
        return

    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json, text/plain, */*",
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
    }

    async with httpx.AsyncClient(timeout=20.0) as client:
        print("Descargando servicios de Nexus 365...")
        services = await fetch_all_services(client, headers)
        print(f"Total servicios descargados: {len(services)}")
        
        print("Consultando detalle de costos e intervenciones...")
        costs_map = await fetch_cost_details(client, headers)

    # Transformación de datos exactos a la plantilla "SERVICIO (2)"
    rows = []
    for idx, item in enumerate(services, start=1):
        code = item.get("code") or item.get("dossierCode") or ""
        cost_info = costs_map.get(code, {})
        
        # Apertura de servicio
        created_at = item.get("createdAt", "")
        if created_at and "T" in str(created_at):
            created_at = str(created_at).split("T")[0]

        # Reprogramaciones
        reprogramado = "SI" if item.get("isRescheduled") or item.get("rescheduledDate") else "NO"
        fecha_reprog = item.get("rescheduledDate", "")
        if fecha_reprog and "T" in str(fecha_reprog):
            fecha_reprog = str(fecha_reprog).split("T")[0]

        # Técnicos
        technicians = item.get("technicians", [])
        if isinstance(technicians, list) and technicians:
            tech_names = ", ".join([t.get("name", "") if isinstance(t, dict) else str(t) for t in technicians])
        else:
            tech_names = item.get("assignedTechnician", {}).get("name", "") or ""

        # Costo total calculado
        total_cost = cost_info.get("totalAmount") or item.get("totalCost") or item.get("amount") or 0

        # Estado del servicio
        status = item.get("currentStatus") or item.get("status") or "ABIERTO"

        rows.append({
            "ITEM": idx,
            "Expediente": code,
            "FECHA APERTURA SERVICIO": created_at,
            "REPROGRAMACIÓN": reprogramado,
            "FECHA REPROGRAMACIÓN": fecha_reprog,
            "DETALLER DEL SERVICIO": item.get("serviceType") or item.get("description") or "",
            "TECNICO QUE HAN INTERVENDO": tech_names,
            "OBSERVACIÓN TECNICO": item.get("technicalNotes") or item.get("comments") or "",
            "COSTO": total_cost,
            "ESTADO": status
        })

    df = pd.DataFrame(rows)
    output_filename = "Seguimiento_Diario_IKE.xlsx"
    
    # Exportación con formato estilizado idéntico al de Luisa
    with pd.ExcelWriter(output_filename, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="SERVICIO (2)", index=False)
        worksheet = writer.sheets["SERVICIO (2)"]

        # Estilos visuales
        header_fill = PatternFill(start_color="0B2545", end_color="0B2545", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        border_side = Side(border_style="thin", color="D3D3D3")
        grid_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

        for col_num, col_name in enumerate(df.columns, 1):
            cell = worksheet.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = grid_border
                cell.alignment = Alignment(vertical="center")

        # Autoajuste de ancho de columnas
        for col in worksheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = col[0].column_letter
            worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

    print(f"\n¡Éxito! Archivo generado correctamente: '{output_filename}' con {len(rows)} registros.")

if __name__ == "__main__":
    asyncio.run(generate_seguimiento_excel())